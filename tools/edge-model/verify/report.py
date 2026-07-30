#!/usr/bin/env python3
"""Diff on-device scores against TANUH's xlsx; emit evidence + PASS/FAIL.
Uses only stdlib + openpyxl. Reads fixtures from $TANUH_FIXTURES (external to the repo).

Pass condition is VERDICT parity — every image's refer / no-refer call must match the reference.
TANUH settled this on 2026-07-17 ("TANUH AI on Avni" §6, Resolved): the per-model sigmoid band
below cannot hold on a device, because Android's Skia JPEG decode differs by ±1 grey level from
desktop libjpeg/OpenCV and MobileViT-v2 amplifies that into sigmoid swings up to ~0.4 on borderline
images. It never flips a verdict, unanimous-AND is itself the borderline protection, and ±1-LSB
robustness is TANUH's to close in model development. Gating on the band would fail every good build.

The per-model table is still computed and printed as evidence, with the 1e-2 band shown for
reference — it is informational and must not affect the exit code."""
import csv, os, sys
from openpyxl import load_workbook

# Reference band only (avg 1e-7..1e-5, max 1e-4..1e-2). Reported, never gating — see module docstring.
MAX_BAR = 1e-2
FOLDS = ["model6", "model8", "model8-2"]
XLSX_COL = {"model6": "Probability_model6", "model8": "Probability_model8", "model8-2": "Probability_model8-2"}


def load_ref(xlsx):
    wb = load_workbook(xlsx, read_only=True); ws = wb.active
    rows = list(ws.iter_rows(values_only=True)); hdr = list(rows[0]); idx = {h: i for i, h in enumerate(hdr)}
    id_col = idx["Image ID"]
    ref = {}; incomplete = []
    for r in rows[1:]:
        # xlsx has trailing/blank rows — skip anything without an Image ID cell.
        if not r or len(r) <= id_col or r[id_col] is None:
            continue
        iid = str(r[id_col]).strip()
        try:
            probs = {f: float(r[idx[XLSX_COL[f]]]) for f in FOLDS}
        except (TypeError, ValueError):
            incomplete.append(iid); continue   # blank/non-numeric reference score — can't compare this image
        ref[iid] = {"true": r[idx["True Label"]], **probs}
    return ref, incomplete


def main():
    fx = os.environ["TANUH_FIXTURES"]
    xlsx = os.path.join(fx, "tanuh_test_data", "Data_models_protocol_for_testing_AI_model_integrations",
                        "Test data", "true_label_model_output_probabilities.xlsx")
    here = os.path.dirname(os.path.abspath(__file__)); out = os.path.join(here, "out")
    dev = {}
    with open(os.path.join(out, "per_model_scores.csv")) as fh:
        for row in csv.DictReader(fh):
            dev[row["image_id"]] = {f: float(row[f]) for f in FOLDS}
    ref, incomplete = load_ref(xlsx)

    joined = [iid for iid in dev if iid in ref]
    dev_incomplete = sorted(iid for iid in dev if iid in set(incomplete))
    if not joined:
        print(f"❌ 0 images joined (dev {len(dev)}, xlsx {len(ref)}) — image IDs don't match the xlsx. "
              f"Check the pushed image basenames vs the 'Image ID' column.", file=sys.stderr)
        sys.exit(1)
    per = {f: {"max": 0.0, "sum": 0.0, "ge": 0} for f in FOLDS}
    rows_out = []
    verdict_match = 0
    for iid in sorted(joined):
        d, r = dev[iid], ref[iid]
        and_dev = all(d[f] > 0.5 for f in FOLDS); and_ref = all(r[f] > 0.5 for f in FOLDS)
        verdict_match += int(and_dev == and_ref)
        rec = {"image_id": iid, "true": r["true"], "dev_verdict": int(and_dev), "ref_verdict": int(and_ref)}
        for f in FOLDS:
            diff = abs(d[f] - r[f]); per[f]["max"] = max(per[f]["max"], diff)
            per[f]["sum"] += diff; per[f]["ge"] += int(diff >= MAX_BAR)
            rec[f"{f}_dev"], rec[f"{f}_ref"], rec[f"{f}_diff"] = d[f], r[f], diff
        rows_out.append(rec)

    n = len(joined)
    worst = max(per[f]["max"] for f in FOLDS); total_ge = sum(per[f]["ge"] for f in FOLDS)
    ok = verdict_match == n          # verdict parity is the pass condition (TANUH, 2026-07-17)
    with open(os.path.join(out, "per_image_scores.csv"), "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows_out[0].keys())); w.writeheader(); w.writerows(rows_out)
    mismatched = [r["image_id"] for r in rows_out if r["dev_verdict"] != r["ref_verdict"]]
    lines = [f"# On-device parity — avni corrected pipeline vs TANUH xlsx", "",
             f"- images joined: **{n}** (dev {len(dev)}, xlsx-with-complete-scores {len(ref)})",
             f"- skipped (reference incomplete in xlsx): {len(dev_incomplete)}"
             + (f" — {', '.join(dev_incomplete)}" if dev_incomplete else ""),
             "- acceptance bar: **verdict-level parity** — every image's refer / no-refer call must",
             "  match the reference (TANUH, 2026-07-17; \"TANUH AI on Avni\" §6, Resolved)", "",
             f"**Verdict parity vs reference: {verdict_match}/{n} — {'PASS ✅' if ok else 'FAIL ❌'}**"]
    if mismatched:
        lines += ["", f"- verdict mismatches: {', '.join(mismatched)}"]
    lines += ["", "## Per-model sigmoid difference (evidence only — does not gate this run)", "",
              f"Reference band for context: max |sigmoid diff| < {MAX_BAR:g}. Exceeding it on a device is",
              "expected: Android's Skia JPEG decode differs by ±1 grey level from desktop libjpeg/OpenCV",
              "and MobileViT-v2 amplifies that on borderline images, without changing a verdict.", "",
              f"| model | max diff | mean diff | # images ≥ {MAX_BAR:g} |", "|---|---|---|---|"]
    for f in FOLDS:
        lines.append(f"| {f} | {per[f]['max']:.3e} | {per[f]['sum']/n:.3e} | {per[f]['ge']} |")
    lines += ["", f"worst per-model diff = {worst:.3e}; images at or above the band = {total_ge}"]
    open(os.path.join(out, "summary.md"), "w").write("\n".join(lines) + "\n")
    print("\n".join(lines))
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
